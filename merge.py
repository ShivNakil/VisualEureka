import os
import cv2
import pytesseract
import openpyxl
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from PIL import Image
from tkinter import Tk, filedialog
from tqdm import tqdm

# Set the Tesseract executable path (change this according to your configuration)
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

# Open a file dialog for the user to select a folder containing images
Tk().withdraw()  # Hide the main tkinter window
folder_path = filedialog.askdirectory(title="Select a folder containing images")

# Check if a folder was selected
if folder_path:
    # Create an Excel workbook and add a worksheet
    excel_file_path = os.path.join(folder_path, "combined_extraction.xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Combined Extraction"

    # Add headers to the worksheet
    worksheet["A1"] = "Image Name"
    worksheet["B1"] = "Extracted Text (Method 1)"
    worksheet["C1"] = "Extracted Text (Method 2)"

    # Initialize the TrOCR model
    processor = TrOCRProcessor.from_pretrained('microsoft/trocr-base-handwritten')
    model = VisionEncoderDecoderModel.from_pretrained('microsoft/trocr-base-handwritten')
    

    # Get a list of image files in the selected folder
    image_files = [filename for filename in os.listdir(folder_path) if filename.endswith((".png", ".jpg", ".jpeg"))]

    # Create a tqdm progress bar
    progress_bar = tqdm(image_files, unit="image")

    # Iterate over the files in the selected folder
    for filename in progress_bar:
        # Check if the file is an image
        if filename.endswith((".png", ".jpg", ".jpeg")):
            # Construct the full path to the image file
            image_path = os.path.join(folder_path, filename)

            # Method 1: Text extraction using TrOCR
            image_trocr = Image.open(image_path).convert("RGB")
            pixel_values = processor(images=image_trocr, return_tensors="pt").pixel_values
            generated_ids = model.generate(pixel_values, max_length=200)
            extracted_text_trocr = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

            # Method 2: Text extraction using Tesseract
            image_cv2 = cv2.imread(image_path)
            gray_image = cv2.cvtColor(image_cv2, cv2.COLOR_BGR2GRAY)
            extracted_text_tesseract = pytesseract.image_to_string(gray_image)

            # Add the image name and extracted text to the Excel worksheet
            row_number = worksheet.max_row + 1
            worksheet.cell(row=row_number, column=1, value=filename)
            worksheet.cell(row=row_number, column=2, value=extracted_text_trocr)
            worksheet.cell(row=row_number, column=3, value=extracted_text_tesseract)

    # Close the tqdm progress bar
    progress_bar.close()

    # Save the Excel file
    workbook.save(excel_file_path)

    print("Extraction complete. Combined results saved in:", excel_file_path)
else:
    print("No folder selected.")