import pytesseract
import cv2
import os
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askdirectory

# Path to the Tesseract executable (change this according to your configuration)
pytesseract.pytesseract.tesseract_cmd = r"C:/Program Files/Tesseract-OCR/tesseract.exe"

# Open a file dialog for the user to select a folder containing images
Tk().withdraw()  # Hide the main tkinter window
folder_path = askdirectory(title="Select a folder containing images")

# Check if a folder was selected
if folder_path:
    # Create an Excel workbook and add a worksheet
    excel_file_path = os.path.join(folder_path, "extracted_text.xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Extracted Text"

    # Add headers to the worksheet
    worksheet["A1"] = "Image Name"
    worksheet["B1"] = "Extracted Text"

    # Iterate over the files in the selected folder
    for filename in os.listdir(folder_path):
        # Check if the file is an image
        if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith(".jpeg"):
            # Construct the full path to the image file
            image_path = os.path.join(folder_path, filename)

            # Open the image using OpenCV
            image = cv2.imread(image_path)

            # Convert the image to grayscale
            gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

            # Use Tesseract to extract text from the image
            extracted_text = pytesseract.image_to_string(gray_image)

            # Add the image name and extracted text to the Excel worksheet
            row_number = worksheet.max_row + 1
            worksheet.cell(row=row_number, column=1, value=filename)
            worksheet.cell(row=row_number, column=2, value=extracted_text)

    # Save the Excel file
    workbook.save(excel_file_path)

    print("Extraction complete. Extracted text is saved in:", excel_file_path)
else:
    print("No folder selected.")
